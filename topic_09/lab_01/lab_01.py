# Create a Python module to create the following Excel spreadsheet (shown below) using
# openpyxl module. Use a random number generator (20-100) to generate the sales in each cell

import random
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Define the product categories
products = [
    "Tablets",
    "Phones",
    "Furniture",
    "Stationary",
    "Printers",
    "Laptops",
    "Desktops"
]

workbook = Workbook()
active_sheet = workbook.active

# Define border style
border_style = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Set column headers and apply border
active_sheet["A1"] = "Product"
for column in range(2, 5):
    cell = active_sheet.cell(row=1, column=column)
    cell.value = f"Store {column-1}"
    cell.border = border_style
    cell.alignment = Alignment(horizontal='center')

# Set products
for row in range(2, len(products) + 2):
    cell = active_sheet.cell(row=row, column=1)
    cell.value = products[row-2]
    cell.border = border_style

# Set random sales values, apply border and centre
for column in range(2, 5):
    for row in range(2, len(products) + 2):
        cell = active_sheet.cell(row=row, column=column)
        cell.value = random.randint(20, 100)
        cell.border = border_style
        cell.alignment = Alignment(horizontal='center')

# Resize product column
max_product_length = max(len(product) for product in products)
active_sheet.column_dimensions["A"].width = max_product_length + 2

workbook.save(filename="lab_01.xlsx")
