from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# Open the workbook and set the active sheet
workbook = load_workbook("lab_03.xlsx")
sheet = workbook.active

# Get row and column count
row_count = sheet.max_row
col_count = sheet.max_column

# Create the Bar Chart
chart = BarChart()

# Set the data and titles range
data = Reference(worksheet=sheet,
                 min_row=1,
                 max_row=row_count,
                 min_col=2,
                 max_col=col_count
                 )

# Set the categories range
categories = Reference(worksheet=sheet,
                       min_row=2,
                       max_row=row_count,
                       min_col=1,
                       max_col=1)

chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)
sheet.add_chart(chart, "A10")

workbook.save("lab_03.xlsx")

