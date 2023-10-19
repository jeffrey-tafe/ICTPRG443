from openpyxl import load_workbook


my_workbook = load_workbook('products.xlsx')
active_sheet = my_workbook['Reviews']

for my_value in active_sheet.iter_rows(min_row=1, max_row=5, min_col=1, max_col=3, values_only=True):
    print(my_value)

print("\n")

for my_value in active_sheet.iter_rows(min_row=1, max_row=5, min_col=1, max_col=3, values_only=True):
    print(f"{my_value[0]}\t{my_value[1]}\t{my_value[2]}")

print("\n")

for my_value in active_sheet.iter_rows(values_only=True):
    print(f"{my_value[0]}\t{my_value[1]}\t{my_value[2]}")

for my_value in active_sheet.iter_rows(values_only=True):
    print(f"{my_value}")

print("\n")
print("\nBy Columns")

for my_value in active_sheet.iter_cols(values_only=True):
    print(f"{my_value}")