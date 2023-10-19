import openpyxl
from openpyxl.styles import Font, PatternFill, Color, GradientFill, Alignment, Border, Side

# Create a new workbook
my_workbook = openpyxl.Workbook()
active_sheet= my_workbook.active

active_sheet['A1'] = "Hello"
active_sheet['B1'] = "World"
active_sheet.title = "Hello_Sheet"

new_sheet_name = "World_sheet"
my_workbook.create_sheet(new_sheet_name)
active_sheet = my_workbook[new_sheet_name]

active_sheet['A1'] = "Welcome to"
active_sheet['A2'] = "Python openpyxl"
active_sheet = my_workbook["Hello_Sheet"]
active_sheet['C1'] = "How are you?"

active_sheet.sheet_properties.TabColor = "FFFF00"
active_sheet['C1'].font = Font(bold=True)
active_sheet['C1'].font = Font(bold=True, italic=True, underline='double', color='0000FF')

# openpyxl.readthedocs.io/en/latest/styles.html

# Set Row height and column width
active_sheet.column_dimensions['A'].width = 20
active_sheet.row_dimensions[1].height = 40

# Background colour
# Need to import:  from openpyxl.styles import PatternFill
active_sheet['C1'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type= 'solid')
# Gradient Fill

# Need to import:  from openpyxl.styles import GradientFill, Color
start_color = Color(rgb='FF0000') # Red
end_color = Color(rgb='00FF00') # Green
active_sheet['C1'].fill = GradientFill(stop=[start_color, end_color])
# Alignment
active_sheet['C1'].alignment = Alignment(horizontal='center', vertical='center') # Need to import:  from openpyxl.styles import Alignment
# Borders, Need to import:  from openpyxl.styles import Border, Side
thin = Side(border_style="thin", color='FF00FF') #Border styles can be dashDot’,‘dashDotDot’,‘dashed’,‘dotted’,‘double’,‘hair’,‘medium’,‘mediumDashDot’,‘mediumDashDotDot’,,‘mediumDashed’,‘slantDashDot’,‘thick’,‘thin’
double = Side(border_style="double", color='008000')
dash_dot = Side(border_style="dashDot", color='0000FF')
slant_dash_dot = Side(border_style="slantDashDot", color='0000FF')
active_sheet['C1'].border = Border(top=thin, left=dash_dot, right=double, bottom=slant_dash_dot)
active_sheet['A2'] = 5
active_sheet['A3'] = 6
active_sheet['A4'] = 7
total = 0
cell_range = 'A2:A4'
for row in active_sheet[cell_range]:
    for cell in row:
        total = total + cell.value
print(f"Total for {cell_range} is {total}")
try:
    my_workbook.save(filename='hello_world.xlsx')
    print("Workbook saved!")
except PermissionError:
    print("Error: Cannot save file at this time, file may be in use")
except Exception as ex:
    print(f"{type(ex).__name__, ex.args}")