from openpyxl import load_workbook

FILE_NAME = 'hello_world.xlsx'

my_workbook = load_workbook(filename=FILE_NAME)

active_sheet = my_workbook['Hello_Sheet']

print(active_sheet['A1'].value)
print(active_sheet.cell(row=1,column=2).value)

print(f"Sheet: {active_sheet}")
print(f"Sheet Title: {active_sheet.title}")
print(f"Sheet Names: {my_workbook.sheetnames}")

print(f"Sheet Count: {len(my_workbook.sheetnames)}")

try:
    my_workbook.save(filename='hello_world.xlsx')
    print("Workbook saved!")
except PermissionError:
    print("Error: Cannot save file at this time, file may be in use")
except Exception as ex:
    print(f"{type(ex).__name__, ex.args}")