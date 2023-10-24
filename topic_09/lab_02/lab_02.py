from student import Student
from openpyxl import load_workbook

# Create empty student list
students = []

# Open the workbook and set the active sheet
workbook = load_workbook(filename='students.xlsx')
sheet = workbook.active

# For each row, skipping headers
for row in sheet.iter_rows(min_row=2):

    # Create a student object
    student = Student(row[0].value, row[1].value, row[2].value, row[3].value)

    # Add student object to students list
    students.append(student)

# Print each student in students list
for student in students:
    print(f"{student.id} - {student.name}")
