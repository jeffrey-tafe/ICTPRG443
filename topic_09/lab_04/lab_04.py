from openpyxl import load_workbook, Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Border, Side
import os

# Delete existing result workbook if exists
result_workbook_path = "result.xlsx"
if os.path.exists(result_workbook_path):
    os.remove(result_workbook_path)

# Open source workbook and set source sheet
source_workbook = load_workbook("covid.xlsx")
source_sheet = source_workbook.active

# Create result workbook and set result sheet
result_workbook = Workbook()
result_sheet = result_workbook.active

# Get row and column count
row_count = source_sheet.max_row
col_count = source_sheet.max_column

# Set border style
border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Copy data from source worksheet to result worksheet
for row in source_sheet.iter_rows(values_only=True):
    result_sheet.append(row)

# Set border style on each cell in result worksheet
for row in result_sheet.iter_rows():
    for cell in row:
        cell.border = border_style

# Create data and titles range
data = Reference(worksheet=result_sheet,
                 min_row=2,
                 max_row=row_count,
                 min_col=1,
                 max_col=col_count)

# Create categories range
categories = Reference(worksheet=result_sheet,
                       min_row=1,
                       max_row=1,
                       min_col=2,
                       max_col=col_count)

# Create chart
chart = LineChart()
chart.add_data(data,
               from_rows=True,
               titles_from_data=True)
chart.set_categories(categories)

result_sheet.add_chart(chart, "A10")
result_workbook.save("result.xlsx")

